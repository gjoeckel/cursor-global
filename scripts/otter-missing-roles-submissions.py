#!/usr/bin/env python3
"""
Process OTTER_missing_roles_submissions.xlsx:
1. Standardize first_name (C), last_name (D), email (E): trim + lowercase (row 2+) both sheets
2. Dedupe Submissions by first_name+last_name+email (keep first)
3-5. Email lookup from Submissions -> Missing_Roles; create Missing_Roles_Email
6-8. Name lookup for blank role/org; create Missing_Roles_Email_Names with "nomatch" for remaining blanks
Uses in-memory data for speed on large sheets.
"""
import openpyxl
from pathlib import Path

PATH = Path("/Users/a00288946/Agents/resources/otter/OTTER_missing_roles_submissions.xlsx")
C_FNAME, C_LNAME, C_EMAIL = 3, 4, 5
C_ROLE, C_ORG = 10, 11


def _norm(v):
    if v is None or not isinstance(v, str):
        return v
    return v.strip().lower()


def _blank(v):
    return v is None or (isinstance(v, str) and not v.strip())


def sheet_to_data(ws, max_col):
    return [[ws.cell(row=r, column=c).value for c in range(1, max_col + 1)] for r in range(1, ws.max_row + 1)]


def data_to_sheet(ws, data, max_col):
    for r, row in enumerate(data, start=1):
        for c in range(1, max_col + 1):
            val = row[c - 1] if c <= len(row) else None
            ws.cell(row=r, column=c).value = val


def main():
    wb = openpyxl.load_workbook(PATH, data_only=True)
    mr_ws = wb["Missing_Roles"]
    sub_ws = wb["Submissions"]
    mr_data = [[mr_ws.cell(row=r, column=c).value for c in range(1, 12)] for r in range(1, mr_ws.max_row + 1)]
    sub_data = [[sub_ws.cell(row=r, column=c).value for c in range(1, 12)] for r in range(1, sub_ws.max_row + 1)]

    max_col = 11

    # 1. Standardize (row 2+) first_name, last_name, email
    for rows in (mr_data, sub_data):
        for row in rows[1:]:
            for i in (C_FNAME - 1, C_LNAME - 1, C_EMAIL - 1):
                v = row[i]
                row[i] = _norm(v) if v is not None else v

    # 2. Dedupe Submissions (keep first)
    seen = set()
    keep = [sub_data[0]]
    for row in sub_data[1:]:
        key = (row[C_FNAME - 1], row[C_LNAME - 1], row[C_EMAIL - 1])
        if key not in seen:
            seen.add(key)
            keep.append(row)
    sub_data = keep

    # Build email -> (role, org) from Submissions
    email_to = {}
    for row in sub_data[1:]:
        em = row[C_EMAIL - 1]
        if not _blank(em) and em not in email_to:
            email_to[em] = (row[C_ROLE - 1], row[C_ORG - 1])

    # 3-5. Missing_Roles_Email: Missing_Roles + role/org from email match
    mr_email_data = []
    for row in mr_data:
        new_row = list(row)
        if len(mr_email_data) >= 1:  # data row
            em = new_row[C_EMAIL - 1]
            if em and em in email_to:
                role, org = email_to[em]
                new_row[C_ROLE - 1] = role
                new_row[C_ORG - 1] = org
        mr_email_data.append(new_row)

    # Build (first_name, last_name) -> (role, org) from Submissions
    name_to = {}
    for row in sub_data[1:]:
        fn, ln = row[C_FNAME - 1], row[C_LNAME - 1]
        if fn is not None or ln is not None:
            key = (fn, ln)
            if key not in name_to:
                name_to[key] = (row[C_ROLE - 1], row[C_ORG - 1])

    # 6-7. Fill blank role/org in Missing_Roles_Email from name lookup
    for row in mr_email_data[1:]:
        if _blank(row[C_ROLE - 1]) and _blank(row[C_ORG - 1]):
            key = (row[C_FNAME - 1], row[C_LNAME - 1])
            if key in name_to:
                row[C_ROLE - 1], row[C_ORG - 1] = name_to[key]

    # 8. Missing_Roles_Email_Names: same + "nomatch" for any remaining blank role/org
    mr_names_data = [list(row) for row in mr_email_data]
    for row in mr_names_data[1:]:
        if _blank(row[C_ROLE - 1]):
            row[C_ROLE - 1] = "nomatch"
        if _blank(row[C_ORG - 1]):
            row[C_ORG - 1] = "nomatch"

    # Write back: load workbook in write mode
    wb = openpyxl.load_workbook(PATH)
    mr_ws = wb["Missing_Roles"]
    sub_ws = wb["Submissions"]

    data_to_sheet(mr_ws, mr_data, max_col)
    data_to_sheet(sub_ws, sub_data, max_col)

    ws_email = wb.create_sheet("Missing_Roles_Email", 2)
    data_to_sheet(ws_email, mr_email_data, max_col)
    ws_names = wb.create_sheet("Missing_Roles_Email_Names", 3)
    data_to_sheet(ws_names, mr_names_data, max_col)

    wb.save(PATH)
    print("Done. Missing_Roles_Email and Missing_Roles_Email_Names created; Submissions deduped; data standardized.")


if __name__ == "__main__":
    main()
